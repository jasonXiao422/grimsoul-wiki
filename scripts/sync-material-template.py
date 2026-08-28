#!/usr/bin/env python3
"""把 materials.json 中新增的材料追加到材料介绍 Excel 模板。"""
import json
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MATERIALS_PATH = PROJECT_ROOT / "src" / "data" / "materials.json"
TEMPLATE_PATH = PROJECT_ROOT / "data-source" / "材料数据.xlsx"
SHEET_NAME = "材料介绍"


# 与 import_excel.py 保持一致：跨表材料译名统一到右侧名称。
MATERIAL_ALIAS = {
    "绳索": "绳子",
    "金属线": "线",
    "附魔皮革": "魔法皮革",
    "真银": "纯银",
    "肖博尔之烬": "修博尔的灰烬",
    "肖博尔骨灰": "修博尔的灰烬",
    "古代板甲": "古老薄板",
    "铜碎片": "碎铜片",
}


def normalize_material_name(value):
    """按 import_excel.py 的既有规则清理名称并归一化别名。"""
    if value is None:
        return None
    name = str(value).replace("\u3000", " ").strip()
    if not name:
        return None
    name = re.sub(r"\s+", "", name)
    return MATERIAL_ALIAS.get(name, name)


def copy_cell_style(source_cell, target_cell):
    """复制现有单元格格式，但不共享可变样式对象。"""
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def main() -> None:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"未找到材料模板：{TEMPLATE_PATH}\n"
            "请先运行 python scripts/gen-material-template.py。"
        )

    with MATERIALS_PATH.open("r", encoding="utf-8") as source_file:
        materials = json.load(source_file)

    workbook = load_workbook(TEMPLATE_PATH)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"Excel 中缺少工作表「{SHEET_NAME}」：{TEMPLATE_PATH}")
    worksheet = workbook[SHEET_NAME]

    existing_names = set()
    for row in worksheet.iter_rows(min_col=1, max_col=1):
        name = normalize_material_name(row[0].value)
        if name:
            existing_names.add(name)

    new_names = []
    seen_json_names = set()
    for material in materials:
        name = material.get("name")
        normalized_name = normalize_material_name(name)
        if not normalized_name or normalized_name in existing_names:
            continue
        if normalized_name in seen_json_names:
            continue
        seen_json_names.add(normalized_name)
        new_names.append(str(name).replace("\u3000", " ").strip())

    if not new_names:
        print("材料 Excel 已包含 materials.json 中的全部材料，没有需要追加的新材料。")
        return

    template_row = worksheet.max_row if worksheet.max_row >= 1 else None
    for name in new_names:
        row_number = worksheet.max_row + 1
        name_cell = worksheet.cell(row=row_number, column=1, value=name)
        description_cell = worksheet.cell(row=row_number, column=2, value=None)

        if template_row:
            copy_cell_style(worksheet.cell(row=template_row, column=1), name_cell)
            copy_cell_style(worksheet.cell(row=template_row, column=2), description_cell)

        # 品质由用户后续手动标注，新追加行不带任何填充色。
        name_cell.fill = PatternFill(fill_type=None)
        description_cell.fill = PatternFill(fill_type=None)
        if template_row and worksheet.row_dimensions[template_row].height is not None:
            worksheet.row_dimensions[row_number].height = worksheet.row_dimensions[template_row].height

    workbook.save(TEMPLATE_PATH)
    print(f"已追加材料：{', '.join(new_names)}")
    print(f"追加总条数：{len(new_names)}")


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, ValueError, OSError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
