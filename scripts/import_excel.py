#!/usr/bin/env python3
"""
把 data-source/ 下的 Excel 转成 src/data/ 下的 JSON。

以后拿到新版本数据时：
  1. 用新 Excel 覆盖 data-source/ 里的同名文件（或改下面的 FILES 常量）
  2. 运行  python3 scripts/import_excel.py
  3. 运行  npm run build  确认通过

脚本会在结尾打印一份报告，列出解析失败或数据可疑的条目。
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pypinyin import lazy_pinyin

# Windows 控制台默认 GBK 编码，打印中文警告会抛 UnicodeEncodeError，
# 导致数据其实已经写完、进程却以失败退出。强制标准输出用 UTF-8。
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data-source"
OUT = ROOT / "src" / "data"

FILES = {
    "weapons": ("武器数据.xlsx", "Sheet1"),
    "armor": ("护甲数据.xlsx", "护甲"),
    "shields": ("盾牌数据.xlsx", "护甲"),
    "backpacks": ("驮篮数据.xlsx", "Sheet1"),
    "enemies": ("敌人数据.xlsx", "敌人生命护甲和伤害"),
    "amulets": ("护符数据.xlsx", "Sheet1"),
    "scrolls": ("卷轴数据.xlsx", "Sheet1"),
    "runes": ("符文数据.xlsx", "Sheet1"),
    "consumables": ("食物_药数据.xlsx", "Sheet1"),
    "sharpen": ("磨尖武器数据.xlsx", "磨刀等级"),
    "boxes": ("武器盒子数据.xlsx", "Sheet1"),
    "materials": ("材料数据.xlsx", "材料介绍"),
    "cabinets": ("柜子数据.xlsx", "Sheet1"),
    "surface-chests": ("地表箱子数据.xlsx", "Sheet1"),
    "fixed-buildings": ("不可升级建筑数据.xlsx", "Sheet1"),
    "skills": ("技能数据.xlsx", "技能"),
}

warnings = []
used_ids = {}
skill_drop_parse_warnings = 0

# 全站规则：技能的 Lv11 及以上等级统一通过「打牌」获得。
# 该规则不写在 Excel 中，由导入时按技能实际最大等级自动附加。
# 仅对确实拥有 Lv11+ 数据的技能生效。若游戏机制变更，只需修改此处。
HIGH_LEVEL_SOURCE_NAME = "打牌"
HIGH_LEVEL_THRESHOLD = 11
high_level_source_added_count = 0
high_level_source_skipped_count = 0


# ---------------------------------------------------------------- 工具函数

def clean(v):
    """去掉首尾空白和全角空格；空值/占位符统一成 None。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return None  # Excel 把数值误存成日期，原值已丢失
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace("\u3000", " ").strip()
    s = re.sub(r"\s+\n", "\n", s)
    if s in ("", "/", "-", "—"):
        return None
    return s


def text(v):
    c = clean(v)
    return None if c is None else str(c)


DOT_TYPE_MARKERS = (
    ("点燃", "点燃"),
    ("冻伤", "冻伤"),
    ("中毒", "中毒"),
    ("流血", "流血"),
)


def parse_dot_types(effect):
    """从效果原文提取持续伤害类型；只记录类型，不解析数值。"""
    value = text(effect)
    if not value:
        return []
    return [dot_type for marker, dot_type in DOT_TYPE_MARKERS if marker in value]


def strip_ranged_marker(effect):
    """移除 Excel 中仅用于筛选的「远程武器」标记及相邻标点。"""
    value = text(effect)
    if not value:
        return None
    value = re.sub(r"远程武器\s*[:：]?\s*", "", value)
    value = value.strip().strip("；;：:").strip()
    return value or None


# 拼音相同但实为不同物品时，在此固定 id，避免顺序变化导致 id 漂移。
# 例：青铜锭 dìng 与 青铜钉 dīng 去声调后拼音一致。
ID_OVERRIDE = {
    "青铜锭": "qing-tong-ingot",
    "青铜钉": "qing-tong-nail",
}


def make_id(name, prefix=""):
    fixed = ID_OVERRIDE.get(re.sub(r"[（）()【】《》\s]+", "", name))
    if fixed:
        return f"{prefix}-{fixed}" if prefix else fixed

    """中文名 → 拼音 id。重名自动加数字后缀。"""
    base = "-".join(lazy_pinyin(re.sub(r"[（）()【】《》\s]+", "", name)))
    base = re.sub(r"[^a-z0-9]+", "-", base.lower()).strip("-")
    if not base:
        base = "item"
    if prefix:
        base = f"{prefix}-{base}"
    key = base
    n = used_ids.get(base, 0)
    if n:
        key = f"{base}-{n + 1}"
    used_ids[base] = n + 1
    return key


# 各表之间材料译名不一致时，统一到右侧的名称。
# 例如盾牌表写「绳索」、武器表写「绳子」，实为同一材料。
MATERIAL_ALIAS = {
    "绳索": "绳子",
    "金属线": "线",
    "附魔皮革": "魔法皮革",
    "真银": "纯银",
    "肖博尔之烬": "修博尔的灰烬",
    "肖博尔骨灰": "修博尔的灰烬",
    "古代板甲": "古老薄板",
    "铜碎片": "碎铜片",
}


def parse_cost(raw):
    """
    解析两种材料写法：
      武器： 原松木*1 绳子*1 亚麻纤维*1
      其他： 布料：7；亚麻纤维：13；绳索：2
    返回 [{"material": "布料", "qty": 7}, ...]
    """
    s = text(raw)
    if not s:
        return []
    if s in ("无", "待补充"):
        return []
    s = s.replace("\n", " ").replace("：", ":").replace("；", ";")
    items = []

    if ":" in s:  # 冒号写法
        for part in re.split(r"[;,，]", s):
            part = part.strip()
            if not part:
                continue
            m = re.match(r"^(.+?)\s*:\s*([\d.]+)$", part)
            if m:
                items.append({"material": m.group(1).strip(), "qty": float(m.group(2))})
            else:
                warnings.append(f"材料段落无法解析: {part!r}")
    else:  # 星号写法
        # 有些表用空格分隔，有些表用分号分隔；统一成空格后沿用原解析规则。
        star_s = re.sub(r"[;,，]\s*", " ", s)
        for m in re.finditer(r"([^\s*]+)\s*\*\s*([\d.]+)", star_s):
            items.append({"material": m.group(1).strip(), "qty": float(m.group(2))})
        if not items:
            warnings.append(f"材料字符串无法解析: {s!r}")

    merged = {}
    for it in items:
        name = MATERIAL_ALIAS.get(it["material"], it["material"])
        qty = it["qty"]
        merged[name] = merged.get(name, 0) + qty
    return [{"material": k, "qty": int(v) if v == int(v) else v} for k, v in merged.items()]


def num(v):
    """浮点整数转 int，避免 JSON 里出现 5.0 这种。"""
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def format_skill_level_value(cell):
    value = clean(cell.value)
    if value is None:
        return None
    if isinstance(value, (int, float)) and "%" in str(cell.number_format or ""):
        percent = round(value * 100, 10)
        return f"{num(percent)}%"
    return value


def parse_damage(raw):
    """'19+5' → (19, 5)；'40真伤' → (40, None) 且标记真伤；纯数字 → (n, None)"""
    c = clean(raw)
    if c is None:
        return None, None, None
    if isinstance(c, (int, float)):
        return c, None, None
    s = str(c)
    note = None
    if "真伤" in s:
        note = "无视护甲（真实伤害）"
        s = s.replace("真伤", "").strip()
    m = re.match(r"^([\d.]+)\s*\+\s*([\d.]+)$", s)
    if m:
        return num(float(m.group(1))), num(float(m.group(2))), note
    m = re.match(r"^([\d.]+)$", s)
    if m:
        return num(float(m.group(1))), None, note
    return None, None, s  # 多段数值等复杂情况，原样保留


def parse_element(raw):
    """'11 冰伤' / '6火伤' / '13 衰败伤' → {'type': '冰', 'value': 11}"""
    s = text(raw)
    if not s:
        return None
    m = re.match(r"^([\d.]+)\s*(.*?)\s*(?:伤|霜|焰)?$", s.replace(" ", " "))
    if not m:
        return {"raw": s}
    val = float(m.group(1))
    kind = m.group(2).strip() or None
    if kind:
        kind = kind.rstrip("伤霜焰")
        kind = {"冰": "冰", "火": "火", "衰败": "衰败", "毒": "毒"}.get(kind, kind)
    return {"type": kind, "value": int(val) if val == int(val) else val}


def parse_durability(raw):
    """50 → {'value': 50}；'400秒' → {'value': 400, 'unit': '秒'}"""
    c = clean(raw)
    if c is None:
        return None
    if isinstance(c, (int, float)):
        return {"value": int(c) if c == int(c) else c, "unit": "点"}
    m = re.match(r"^([\d.]+)\s*(.*)$", str(c))
    if m:
        v = float(m.group(1))
        return {"value": int(v) if v == int(v) else v, "unit": m.group(2).strip() or "点"}
    return {"raw": str(c)}


def parse_formula(raw, label):
    """Excel 公式泄漏成字符串时，尝试算出数值。"""
    c = clean(raw)
    if isinstance(c, (int, float)):
        return c
    s = str(c) if c else ""
    if s.startswith("="):
        m = re.match(r"^=([\d+\-*/. ]+)$", s)
        if m:
            try:
                val = eval(m.group(1))  # 仅含数字和运算符
                warnings.append(f"{label} 的数值是 Excel 公式 {s}，已计算为 {val}，请核对")
                return val
            except Exception:
                pass
        warnings.append(f"{label} 的数值是 Excel 公式 {s}，无法自动计算，已留空")
        return None
    return None



# ---------------------------------------------------------------- 护甲减伤

ARMOR_K = 165  # 减伤 = 护甲 / (护甲 + K)


def undate(cell):
    """
    还原被 Excel 误存为日期的多段数值。

    起因：像「10, 30, 40」这种用逗号分隔的多段数值，Excel 会把它识别成日期，
    存为 1940-10-30，同时把单元格格式设为 m, d, yy。
    显示出来仍然是「10, 30, 40」，肉眼看不出问题，但读取时拿到的是 datetime。

    只要格式里同时含有 m、d、y，就能按 月, 日, 两位年 反推回原始文本。
    """
    if cell is None or not isinstance(cell.value, datetime):
        return None
    fmt = (cell.number_format or "").lower()
    if not ("m" in fmt and "d" in fmt and "y" in fmt):
        return None
    d = cell.value
    return f"{d.month}, {d.day}, {d.year % 100}"


def dr_from_armor(armor):
    """
    由护甲值算减伤百分比，向上取整。
    减伤 = 护甲 / (护甲 + 165)

    支持三种输入：
      120                → 43
      "35, 50, 250"      → "18, 24, 61"     多段（普通/英雄/传奇）
      "不定; <400"        → "不定; <71"       带前缀的上限值
      None               → None
    """
    import math

    def one(v):
        v = float(v)
        return math.ceil(v / (v + ARMOR_K) * 100)

    if armor is None:
        return None
    if isinstance(armor, (int, float)):
        return one(armor)

    s = str(armor).strip()
    if not s:
        return None

    # 提取所有数字，保留原有的分隔与前缀结构
    import re
    nums = re.findall(r"\d+(?:\.\d+)?", s)
    if not nums:
        return s  # 纯文字说明，原样保留

    out = s
    for n in sorted(set(nums), key=len, reverse=True):
        out = out.replace(n, str(one(n)))
    return out


# ---------------------------------------------------------------- 各表解析

def rows_of(key):
    fn, sheet = FILES[key]
    wb = load_workbook(SRC / fn, read_only=True, data_only=False)
    return list(wb[sheet].iter_rows(values_only=True))


def cells_of(key):
    """返回单元格对象而非纯值，用于读取 number_format。"""
    fn, sheet = FILES[key]
    wb = load_workbook(SRC / fn, read_only=True, data_only=False)
    return list(wb[sheet].iter_rows())


def merged_ranges_of(key, column):
    """
    取指定列上的合并单元格区间，返回 [(起始行, 结束行, 左上角的值), ...]。

    cells_of() 用的是 read_only 模式，那种模式下 worksheet 没有 merged_cells，
    所以这里单独再开一次工作簿。一次导入只调用一次，开销可以接受。
    """
    fn, sheet = FILES[key]
    wb = load_workbook(SRC / fn, read_only=False, data_only=False)
    ws = wb[sheet]
    out = []
    for rng in ws.merged_cells.ranges:
        if rng.min_col != column or rng.min_row == 1:
            continue
        out.append((rng.min_row, rng.max_row, ws.cell(row=rng.min_row, column=rng.min_col).value))
    wb.close()
    return out


def build_weapons():
    out = []
    for cells in cells_of("weapons")[2:]:
        row = [c.value for c in cells]
        name = text(row[0])
        if not name:
            continue
        # Excel 里手动 Alt+Enter 的换行会带进名称，统一压成空格
        name = re.sub(r"\s*\n\s*", "", name)
        base, skill, dmg_note = parse_damage(row[1])
        effect = text(row[7])
        is_ranged = bool(effect and "远程武器" in effect)
        blueprint = None
        if effect and "图纸" in effect:
            m = re.match(r"^(.+?)的?高级图纸$", effect.strip())
            if m:
                blueprint = m.group(1).strip()
                effect = None
        out.append({
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "damage": base,
            "skillBonus": skill,
            "damageNote": dmg_note,
            "element": parse_element(row[2]),
            "attackSpeed": clean(row[3]),
            "range": clean(row[4]),
            "durability": parse_durability(row[5]),
            "cost": parse_cost(row[6]),
            "effect": strip_ranged_marker(effect),
            "dotTypes": parse_dot_types(effect),
            "isRanged": is_ranged,
            "upgradeOf": blueprint,
        })
    # 把"XX的高级图纸"换成对应武器 id
    by_name = {w["name"]: w["id"] for w in out}
    for w in out:
        if w["upgradeOf"]:
            target = by_name.get(w["upgradeOf"])
            if target:
                w["upgradeOf"] = target
            else:
                warnings.append(f"武器 {w['name']} 的高级图纸来源 {w['upgradeOf']!r} 找不到对应武器")
                w["upgradeOf"] = None
    return out


def build_armor():
    """套装行后面紧跟 5 件部件，靠名字里的『套装（T…级）』识别。"""
    sets, standalone = [], []
    cur = None
    for cells in cells_of("armor")[1:]:
        row = [c.value for c in cells]
        name = text(row[1])
        if not name:
            continue
        is_set = "套装" in name and "级）" in name
        entry_armor = parse_formula(row[2], f"护甲 {name}") if not isinstance(clean(row[2]), (int, float)) else clean(row[2])
        quality = quality_from_fill(cells[1])
        common = {
            "name": re.sub(r"\s+", "", name),
            "quality": quality,
            "armor": entry_armor,
            "protection": parse_element(row[3]),
            "cost": parse_cost(row[5]),
            "effect": text(row[6]),
        }
        if is_set:
            m = re.match(r"^(.+?)套装（(T[\d+]+)级）$", common["name"])
            cur = {
                "id": make_id(common["name"]),
                "name": common["name"],
                "tier": m.group(2) if m else None,
                "quality": quality,
                "obtain": text(row[7]) if len(row) > 7 else None,
                "totalArmor": common["armor"],
                "protection": common["protection"],
                "durability": parse_durability(row[4]),
                "cost": common["cost"],
                "setEffect": common["effect"],
                "pieces": [],
            }
            sets.append(cur)
        elif cur is not None and len(cur["pieces"]) < 5:
            cur["pieces"].append({
                "id": make_id(common["name"]),
                "name": common["name"],
                "quality": quality,
                "armor": common["armor"],
                "protection": common["protection"],
                "cost": common["cost"],
                "effect": common["effect"],
            })
        else:
            cur = None
            standalone.append({
                "id": make_id(common["name"]),
                "name": common["name"],
                "tier": None,
                "quality": quality,
                "obtain": text(row[7]) if len(row) > 7 else None,
                "armor": common["armor"],
                "protection": common["protection"],
                "durability": parse_durability(row[4]),
                "cost": common["cost"],
                "effect": common["effect"],
            })
    for s in sets:
        if len(s["pieces"]) != 5:
            warnings.append(f"套装 {s['name']} 只解析到 {len(s['pieces'])} 件部件（预期 5 件）")
    return sets, standalone


def build_shields():
    """
    盾牌表列序：
      0 名称 / 1 耐久 / 2 防御值 / 3 制作 / 4 效果 / 5 获取途径
    """
    out = []
    for cells in cells_of("shields")[1:]:
        row = [c.value for c in cells]
        name = text(row[0])
        if not name:
            continue
        out.append({
            "id": make_id(name),
            "name": re.sub(r"\s+", "", name),
            "quality": quality_from_fill(cells[0]),
            "armor": clean(row[2]),
            "durability": parse_durability(row[1]),
            "cost": parse_cost(row[3]),
            "effect": text(row[4]),
            "obtain": text(row[5]),
        })
    return out


def parse_protection(raw):
    """
    驮篮的元素防护写法：「防御寒冷15」「防御火焰40」「防御衰败40」
    → {"type": "寒冷", "value": 15}
    """
    t = text(raw)
    if not t:
        return None
    m = re.match(r"^防御(.+?)\s*([\d.]+)$", t)
    if m:
        v = float(m.group(2))
        return {"type": m.group(1).strip(), "value": int(v) if v == int(v) else v}
    return parse_element(raw)


def build_backpacks():
    """
    驮篮表列序：
      0 名称 / 1 储存栏 / 2 元素伤害保护 / 3 制作配方 / 4 效果 / 5 获取途径
    """
    out = []
    for cells in cells_of("backpacks")[1:]:
        row = [c.value for c in cells]
        name = text(row[0])
        if not name:
            continue
        out.append({
            "id": make_id(name),
            "name": re.sub(r"\s+", "", name),
            "quality": quality_from_fill(cells[0]),
            "slots": clean(row[1]),
            "protection": parse_protection(row[2]),
            "cost": parse_cost(row[3]),
            "effect": text(row[4]),
            "obtain": text(row[5]),
        })
    return out


def build_cabinets():
    """柜子表每个柜子占三行，A/F 列分别是合并的名称与储存类型。"""
    rows = cells_of("cabinets")
    name_ranges = merged_ranges_of("cabinets", 1)
    storage_ranges = merged_ranges_of("cabinets", 6)
    names = {start: (end, value) for start, end, value in name_ranges}
    storage_types = {start: (end, value) for start, end, value in storage_ranges}

    out = []
    for row_number, (end, raw_name) in names.items():
        cells = rows[row_number - 1]
        name = text(raw_name)
        if not name:
            continue
        storage_end, raw_storage = storage_types.get(row_number, (end, None))
        levels = []
        for excel_row in range(row_number, end + 1):
            row = [cell.value for cell in rows[excel_row - 1]]
            levels.append({
                "level": text(row[1]),
                "capacity": clean(row[2]),
                "castlePoints": clean(row[3]),
                "cost": parse_cost(row[4]),
            })
        storage = parse_storage_types(raw_storage)
        out.append({
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "storageTypes": storage,
            "levels": levels,
        })
    return out


def build_surface_chests():
    """地表箱子表每个箱子占三行，A/B 列分别是合并的地点与名称。"""
    rows = cells_of("surface-chests")
    location_ranges = merged_ranges_of("surface-chests", 1)
    name_ranges = merged_ranges_of("surface-chests", 2)
    locations = {start: (end, value) for start, end, value in location_ranges}
    names = {start: (end, value) for start, end, value in name_ranges}

    out = []
    for row_number, (end, raw_name) in names.items():
        name = text(raw_name)
        if not name:
            continue
        _, raw_location = locations.get(row_number, (end, None))
        levels = []
        for excel_row in range(row_number, end + 1):
            row = [cell.value for cell in rows[excel_row - 1]]
            levels.append({
                "level": text(row[2]),
                "capacity": clean(row[3]),
                "cost": parse_cost(row[4]),
            })
        out.append({
            "id": make_id(name),
            "name": name,
            "location": text(raw_location),
            "levels": levels,
        })
    return out


def build_fixed_buildings():
    """不可升级建筑表：每行一个建筑，A 列名称底色表示品质。"""
    out = []
    for cells in cells_of("fixed-buildings")[1:]:
        row = [cell.value for cell in cells]
        name = text(row[0] if len(row) > 0 else None)
        if not name:
            continue
        raw_build_cost = text(row[5] if len(row) > 5 else None)
        raw_assemble_cost = text(row[6] if len(row) > 6 else None)
        item = {
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "castlePoints": clean(row[1] if len(row) > 1 else None),
            "capacity": clean(row[2] if len(row) > 2 else None),
            "maxCount": clean(row[3] if len(row) > 3 else None),
            "purpose": text(row[4] if len(row) > 4 else None),
            "buildCost": [] if raw_build_cost in (None, "无") else parse_cost(raw_build_cost),
            "assembleCost": [] if raw_assemble_cost in (None, "无") else parse_cost(raw_assemble_cost),
            "blueprintSource": text(row[7] if len(row) > 7 else None),
            "tags": [tag.strip() for tag in (text(row[8]) or "").split("；") if tag.strip()],
        }
        if raw_build_cost == "待补充":
            item["buildCostPending"] = True
        if raw_assemble_cost == "待补充":
            item["assembleCostPending"] = True
        out.append(item)
    return out


SKILL_CATEGORY_BY_FILL = {
    "FFFFC000": ("主动技能", "#ffc000"),
    "FFE06666": ("伤害", "#e06666"),
    "FF92D050": ("治疗", "#92d050"),
    "FFD8D4CC": ("闪避", "#d8d4cc"),
    "FF8A8579": ("资源", "#8a8579"),
    "FF00B0F0": ("角色", "#00b0f0"),
    "FFC08FE8": ("特殊", "#c08fe8"),
}


def skill_category_from_fill(cell):
    try:
        fg = cell.fill.fgColor
        rgb = fg.rgb if fg and fg.type == "rgb" else None
    except Exception:
        rgb = None
    if rgb not in SKILL_CATEGORY_BY_FILL:
        warnings.append(f"未登记的技能类别底色 {rgb}（{cell.value}），无法识别类别")
        return None, None
    return SKILL_CATEGORY_BY_FILL[rgb]


def parse_skill_drop_locations(raw):
    value = text(raw)
    if not value:
        return []
    locations = []
    for part in value.split("；"):
        part = part.strip()
        if not part:
            continue
        _, separator, location_text = re.split(r"([：:])", part, maxsplit=1) if re.search(r"[：:]", part) else ("", "", part)
        location_text = location_text if separator else part
        for location in location_text.split("/"):
            location = location.strip().strip("；")
            if location and location not in locations:
                locations.append(location)
    return locations


def parse_skill_drop_locations_by_level(raw):
    global skill_drop_parse_warnings
    value = text(raw)
    if not value:
        return []
    out = []
    for part in value.split("；"):
        part = part.strip()
        if not part:
            continue
        match = re.match(r"^Lv(\d+)(?:-Lv(\d+))?\s*[：:]\s*(.+)$", part)
        if not match:
            skill_drop_parse_warnings += 1
            warnings.append(f"技能掉落地点片段无法解析，已跳过: {part!r}")
            continue
        start = int(match.group(1))
        end = int(match.group(2) or start)
        if end < start:
            skill_drop_parse_warnings += 1
            warnings.append(f"技能掉落等级范围无效，已跳过: {part!r}")
            continue
        locations = [location.strip() for location in match.group(3).split("/") if location.strip()]
        if not locations:
            skill_drop_parse_warnings += 1
            warnings.append(f"技能掉落地点为空，已跳过: {part!r}")
            continue
        out.append({"levels": list(range(start, end + 1)), "locations": list(dict.fromkeys(locations))})
    return out


def build_skills():
    global high_level_source_added_count, high_level_source_skipped_count
    rows = cells_of("skills")
    headers = [index + 1 for index, row in enumerate(rows) if text(row[2].value) == "Lvl 1"]
    out = []
    for header_index, header_row in enumerate(headers):
        end_row = headers[header_index + 1] - 1 if header_index + 1 < len(headers) else len(rows)
        category = text(rows[header_row - 1][1].value)
        for excel_row in range(header_row + 1, end_row + 1):
            row = rows[excel_row - 1]
            name = text(row[1].value)
            if not name:
                continue
            category_from_fill, category_color = skill_category_from_fill(row[1])
            if category_from_fill and category_from_fill != category:
                warnings.append(f"技能类别表头与名称底色不一致（{name}：{category} / {category_from_fill}）")
            levels = []
            for level, cell in enumerate(row[2:17], start=1):
                value = format_skill_level_value(cell)
                if value is not None:
                    levels.append({"level": level, "value": value})
            drop_raw = text(row[17].value)
            drop_locations_by_level = parse_skill_drop_locations_by_level(drop_raw)
            high_levels = [level["level"] for level in levels if level["level"] >= HIGH_LEVEL_THRESHOLD]
            drop_locations = parse_skill_drop_locations(drop_raw)
            if high_levels:
                drop_locations_by_level.append({"levels": high_levels, "locations": [HIGH_LEVEL_SOURCE_NAME]})
                if HIGH_LEVEL_SOURCE_NAME not in drop_locations:
                    drop_locations.append(HIGH_LEVEL_SOURCE_NAME)
                high_level_source_added_count += 1
            else:
                high_level_source_skipped_count += 1
            out.append({
                "id": make_id(name, "skill"),
                "name": name,
                "category": category,
                "categoryColor": category_color,
                "levels": levels,
                "dropLocationsRaw": drop_raw or "",
                "dropLocations": drop_locations,
                "dropLocationsByLevel": drop_locations_by_level,
            })
    return out


def parse_storage_types(raw):
    """按顶层逗号拆分储存类型，保留括号内的完整内容。"""
    value = text(raw)
    if not value:
        return []
    parts, current, depth = [], [], 0
    for char in value:
        if char in "（(":
            depth += 1
        elif char in "）)" and depth:
            depth -= 1
        if char in ",，" and depth == 0:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
        else:
            current.append(char)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


QUALITY_KEY = {"普通": "common", "稀有": "rare", "独特": "unique", "传说": "legendary"}

# 武器与护甲表用单元格底色标记品质。
# 同一档位可能存在几个相近色值（手填时选了不同色板），一并归入同一档。
QUALITY_BY_FILL = {
    None:       "common",     # 无填充
    "00000000": "common",
    "FFFFFFFF": "common",
    # 蓝 —— 稀有
    "FF4A86E8": "rare",
    "FF4285F4": "rare",
    "FF02A5E3": "rare",
    "FF3399FF": "rare",      # 卷轴表用的第二种蓝
    "FF0070C0": "rare",      # 材料表用的蓝
    # 黄橙 —— 独特
    "FFFBBC04": "unique",
    "FFF1C232": "unique",
    "FFFFC000": "unique",
    # 紫 —— 传说
    "FF351C75": "legendary",
    "FF674EA7": "legendary",
    "FF9933FF": "legendary",
    "FF808080": "legendary",  # 盾牌表里黑色炽热惩罚护盾用了灰底
    "FF7030A0": "legendary",  # 材料表用的紫
}


def quality_from_fill(cell):
    """读取单元格填充色，映射为品质 key。未登记的颜色回退为普通并告警。"""
    try:
        fg = cell.fill.fgColor
        rgb = fg.rgb if fg and fg.type == "rgb" else None
    except Exception:
        rgb = None
    if rgb not in QUALITY_BY_FILL:
        warnings.append(f"未登记的品质底色 {rgb}（{cell.value}），已按普通处理")
        return "common"
    return QUALITY_BY_FILL[rgb]


# 食物药剂表里表示「没有」的几种写法
CONSUMABLE_EMPTY = {"-", "无", "—", "/"}

# 「材料名：数量」，数量后面可能跟一个括号注释，如「南瓜子：1（需种植）」
RECIPE_ITEM_RE = re.compile(r"^(?P<name>.+?)[：:]\s*(?P<qty>\d+)\s*(?:[（(](?P<note>.*?)[）)])?$")
# 行尾的产出说明，如「，得到浆果饮料*1」
RECIPE_YIELD_RE = re.compile(r"[，,]\s*(?P<yield>得到.+)$")


def split_crafted_at(raw):
    """
    制作地点拆成数组，供筛选器按「包含」匹配。

    表里写「厨师营地；篝火；火」表示三个地方都能做。以前整串当一个值，
    筛选器里就会冒出「厨师营地；篝火；火」这种复合选项，选了只能筛出
    恰好这么写的条目。拆开后每个地点各成一项，勾「篝火」就能把所有
    能在篝火做的都筛出来。
    """
    value = _consumable_value(raw)
    if value is None:
        return None
    parts = [p.strip() for p in re.split(r"[;；、,，/]", str(value)) if p.strip()]
    parts = [p for p in parts if p not in CONSUMABLE_EMPTY]
    # 去重但保持原顺序
    seen, out = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out or None


def _consumable_value(raw):
    value = text(raw)
    if value is None or value.strip() in CONSUMABLE_EMPTY:
        return None
    return value


def healing_sort_key(raw):
    """
    从治愈列抽出一个数值，供列表页排序用。

    这一列写法不统一：「200」「瞬间100」「瞬间120，240」。
    取所有数字之和当作总治愈量 —— 目前只有僧侣药酒是两段
    （瞬间120 + 持续240 = 360），其余都是单个数字。

    药水类没有治愈量，返回 None，排序时排在最后。
    """
    value = _consumable_value(raw)
    if value is None:
        return None
    numbers = [int(n) for n in re.findall(r"\d+", str(value))]
    return sum(numbers) if numbers else None


def parse_recipe_text(raw, owner=""):
    """
    把配方原文拆成结构化的配方组，供详情页像护甲部件那样分组展示。

    一格里可能有多套配方，用换行分隔。每行的形态：

        曼德拉药酒：1；苦药酒：1                      → 无标签，两种材料
        厨师营地配方：烈酒：1；冬青果：5               → 有标签
        厨师营地配方1：山楂：3，得到浆果饮料*1          → 有标签，带产出说明
        南瓜子：1（需种植）                          → 材料带括号注释
        错误的厨师营地配方                           → 只有标签，没有材料

    返回 [{"label": str|None, "items": [{"name","qty","note"}], "yield": str|None}]

    材料名保持原文，不转成 material id：一格内含多套配方，且原料里有
    「曼德拉药酒」「苦药酒」这类本身就是食物的条目，不在 materials.json 里，
    硬映射会污染材料反推表。
    """
    value = _consumable_value(raw)
    if value is None:
        return None

    recipes = []
    for line in str(value).split("\n"):
        line = line.strip()
        if not line:
            continue

        yield_note = None
        matched = RECIPE_YIELD_RE.search(line)
        if matched:
            yield_note = matched.group("yield").strip()
            line = line[: matched.start()].strip()

        chunks = [c.strip() for c in re.split(r"[；;]", line) if c.strip()]
        if not chunks:
            continue

        label = None
        first = chunks[0]
        colons = len(re.findall(r"[：:]", first))
        if colons == 0:
            # 整行没有「材料：数量」，只是一句说明
            recipes.append({"label": first, "items": [], "yield": yield_note})
            continue
        if colons >= 2:
            label, _, first = first.partition("：") if "：" in first else first.partition(":")
            label = label.strip()
            chunks[0] = first.strip()

        items = []
        for chunk in chunks:
            item = RECIPE_ITEM_RE.match(chunk)
            if not item:
                warnings.append(f"配方 {owner}：无法解析「{chunk}」，已按原文保留")
                items.append({"name": chunk, "qty": None, "note": None})
                continue
            items.append({
                "name": item.group("name").strip(),
                "qty": int(item.group("qty")),
                "note": (item.group("note") or "").strip() or None,
            })

        recipes.append({"label": label, "items": items, "yield": yield_note})

    return recipes or None


def build_consumables():
    """
    食物药剂表列序：
      0 名称 / 1 特殊效果 / 2 治愈 / 3 饱食度 / 4 口渴值 / 5 配方 / 6 制作地点

    第 1 行是标题（A1:G1 合并），第 2 行是表头，从第 3 行开始是数据。

    配方列保持原文不解析。一格里可能有多套配方（厨师营地配方 / 默认配方 /
    配方1 / 配方2），而且原料里有「曼德拉药酒」「苦药酒」这类本身就是食物
    的条目，不在 materials.json 里。硬拆成材料 id 会污染材料反推表，
    所以按整段文本存进 recipeText，详情页按换行分行渲染。

    治愈的写法不统一（「200」「瞬间100」「瞬间120，240」），
    保持字符串原样，不要强转数字。
    """
    out = []
    started = False
    for cells in cells_of("consumables"):
        row = [c.value for c in cells]
        first = text(row[0])

        if not started:                      # 跳过标题行与表头行
            if first == "名称":
                started = True
            continue
        if not first:
            continue

        name = first
        out.append({
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "healing": _consumable_value(row[2]),
            "healingSort": healing_sort_key(row[2]),
            "satiety": _consumable_value(row[3]),
            "thirst": _consumable_value(row[4]),
            "craftedAt": split_crafted_at(row[6]),
            "effect": _consumable_value(row[1]),
            "recipes": parse_recipe_text(row[5], name),
        })
    return out


def build_runes():
    """
    符文表列序：
      0 名称 / 1 特殊效果

    结构与卷轴表相同：两列，品质由名称单元格底色决定。
    目前三条全是独特（黄橙 FFC000），该色值已在 QUALITY_BY_FILL 里登记。

    第一行是表头（『符文』『特殊效果』），跳过。
    """
    out = []
    for cells in cells_of("runes")[1:]:
        row = [c.value for c in cells]
        name = text(row[0])
        if not name:
            continue
        effect = text(row[1])
        if not effect:
            warnings.append(f"符文 {name}：没有特殊效果说明")
        out.append({
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "effect": effect,
        })
    return out


def build_scrolls():
    """
    卷轴表列序：
      0 名称 / 1 特殊效果

    品质同样由名称单元格底色决定。卷轴表用的是自己的一套色板
    （蓝 FF02A5E3 / FF3399FF，黄橙 FFC000，紫 9933FF），
    其中 FF3399FF 是这张表独有的，已登记进 QUALITY_BY_FILL。

    第一行是表头（『卷轴』『特殊效果』），跳过。
    """
    out = []
    for cells in cells_of("scrolls")[1:]:
        row = [c.value for c in cells]
        name = text(row[0])
        if not name:
            continue
        effect = text(row[1])
        if not effect:
            warnings.append(f"卷轴 {name}：没有特殊效果说明")
        out.append({
            "id": make_id(name),
            "name": name,
            "quality": quality_from_fill(cells[0]),
            "effect": effect,
        })
    return out


def build_amulets():
    """
    护符表列序：
      0 名称 / 1 品质 / 2 耐久 / 3 元素伤害保护 / 4 制作配方 / 5 效果
    品质存英文 key，中文名与配色在 src/lib/quality.ts 里映射。
    """
    out = []
    for row in rows_of("amulets")[1:]:
        name = text(row[0])
        if not name:
            continue
        q = text(row[1])
        out.append({
            "id": make_id(name),
            "name": re.sub(r"\s+", "", name),
            "quality": QUALITY_KEY.get(q, "common"),
            "durability": parse_durability(row[2]),
            "protection": parse_protection(row[3]),
            "cost": parse_cost(row[4]),
            "effect": text(row[5]),
        })
    return out


# 敌人元素伤害的写法是「数值 类型」，如「30 冰」「2, 4, 10 火」「待测 衰败」。
# 拆成 {type, value} 后就能复用武器那套 element 渲染与 elementColor()。
# 左边是表里可能出现的写法，右边是站内统一使用的短名。
# 长的写法要排在短的前面，否则「火焰」会先被「火」匹配掉。
ENEMY_ELEMENT_WORDS = (
    ("火焰", "火"),
    ("火", "火"),
    ("寒冷", "冰"),
    ("冰", "冰"),
    ("衰败", "衰败"),
    ("雷电", "雷"),
    ("闪电", "雷"),
    ("雷", "雷"),
)


def parse_enemy_element(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    for word, canonical in ENEMY_ELEMENT_WORDS:
        if s.endswith(word):
            value = s[: -len(word)].strip().rstrip(",，").strip()
            return {"type": canonical, "value": value or None}
    # 没写类型的原样保留，渲染时按无色处理
    return {"type": None, "value": s}


# 场地机制正文里的小标题，形如 【英雄模式的减益】
NOTE_BLOCK_RE = re.compile(r"^【(?P<title>[^】]+)】\s*(?P<body>.*)$")
# 小标题下的编号条目，两种写法都要认：
#   （1）一条命：固定减益，死亡后无法返回地下城捡尸体
#   1.\t良好体魄：玩家最大生命值增加40点
NOTE_ITEM_RE = re.compile(
    r"^(?:[（(](?P<idx1>\d+)[）)]|(?P<idx2>\d+)\s*[.、．])\s*"
    r"(?P<name>[^：:]+)[：:]\s*(?P<desc>.*)$"
)

# 这些小标题下的编号条目会配图标，值是图标所在的板块目录
NOTE_ICON_BLOCKS = {
    "英雄模式的减益": "debuffs",       # 被弃地下城
    "英雄模式玩家减益": "debuffs",      # 大车炮台
    "英雄模式玩家增益": "buffs",
    "英雄模式敌方强化": "enemy-buffs",
}

# 同一段场地机制会被该分组的每只敌人各解析一遍，而 make_id 用全局 used_ids
# 去重，重复调用会得到 xxx-2、xxx-3……所以按名称缓存，只算第一次。
_NOTE_ICON_IDS = {}


def note_icon_id(name):
    if name not in _NOTE_ICON_IDS:
        _NOTE_ICON_IDS[name] = make_id(name)
    return _NOTE_ICON_IDS[name]


def parse_group_note(raw):
    """
    把整段场地机制拆成结构化区块，供详情页分块渲染。

    形态：
        【地牢介绍】被弃地下城一共有三层……
        【英雄模式的减益】英雄模式下的地牢会有4种随机减益……
        （1）一条命：固定减益，死亡后无法返回地下城捡尸体
        （2）瘟疫崛起：固定减益，敌人可以复活

    返回 [{"title", "body", "items": [{"index","name","desc","iconCat","iconId"}]}]

    编号条目归属于它上面最近的那个小标题。只有 NOTE_ICON_BLOCKS 里登记的
    小标题，其条目才带 iconCat/iconId —— 酷吏那边的房间介绍也用编号，
    但不需要图标。
    """
    if not raw:
        return None

    blocks = []
    for line in str(raw).split("\n"):
        line = line.strip()
        if not line:
            continue

        matched = NOTE_BLOCK_RE.match(line)
        if matched:
            blocks.append({
                "title": matched.group("title").strip(),
                "body": matched.group("body").strip() or None,
                "items": [],
            })
            continue

        item = NOTE_ITEM_RE.match(line)
        if item and blocks:
            block = blocks[-1]
            entry = {
                "index": int(item.group("idx1") or item.group("idx2")),
                "name": item.group("name").strip(),
                "desc": item.group("desc").strip(),
            }
            icon_cat = NOTE_ICON_BLOCKS.get(block["title"])
            if icon_cat:
                entry["iconCat"] = icon_cat
                entry["iconId"] = note_icon_id(entry["name"])
            block["items"].append(entry)
            continue

        # 既不是小标题也不是编号条目，接到上一块的正文后面
        if blocks:
            block = blocks[-1]
            block["body"] = f"{block['body']}\n{line}" if block["body"] else line
        else:
            blocks.append({"title": None, "body": line, "items": []})

    return blocks or None


def collect_note_icons(enemies):
    """
    把带图标的场地机制条目按板块汇总，返回 {板块: [条目, ...]}。

    图标工具和缺图检查都按板块读 JSON，没有这份清单就没法分配图标。
    同一段机制会被该分组的每只敌人各解析一遍，所以按 (板块, id) 去重。

    注意有几个名称跨分组重复：被弃地下城和大车炮台都有「全速奔跑」
    「超自然反应」「被感染的空气」，但数值不同。同板块内共用一张图标，
    跨板块（减益 vs 敌方强化）则各算各的。
    """
    buckets = {}
    for enemy in enemies:
        for block in enemy.get("groupNoteBlocks") or []:
            for item in block["items"]:
                cat = item.get("iconCat")
                if not cat:
                    continue
                bucket = buckets.setdefault(cat, {})
                bucket.setdefault(item["iconId"], {
                    "id": item["iconId"],
                    "name": item["name"],
                    "desc": item["desc"],
                    "group": enemy["group"],
                    "quality": "common",
                })
    return {cat: list(items.values()) for cat, items in buckets.items()}


def build_enemies():
    """
    敌人表列序：
      0 名称 / 1 生命 / 2 减伤% / 3 物理伤害 / 4 元素伤害 / 5 敌人介绍 / 6 出现地点

    两点与其他表不同：

    1. 品质同样由名称单元格底色决定，走 quality_from_fill()。
       敌人表用的色值（蓝 FF4A86E8 / 黄橙 FFBC04 / 紫 FF351C75）
       已经在 QUALITY_BY_FILL 里登记过，不必新增。

    2. 场地机制单独占 H 列，整组共用一段，用跨行合并单元格表示。
       出现地点（G 列）也可能被合并（比如大车炮台整组都写「大车炮台」），
       那种是正常地点，不是机制说明。两列各读各的，不再靠跨行数猜测。

    表头行（第二列是『生命』）同时充当地点分组标题。
    """
    def merged_lookup(column):
        """把某列的合并区间做成「行号 → 左上角的值」的查表函数。"""
        ranges = merged_ranges_of("enemies", column)

        def at(row_index):
            for start, end, value in ranges:
                if start <= row_index <= end:
                    return value
            return None

        return at

    note_at = merged_lookup(8)       # H 列：场地机制
    location_at = merged_lookup(7)   # G 列：出现地点，合并时把值补给整组

    out, group = [], None
    for cells in cells_of("enemies")[1:]:
        row = [c.value for c in cells]
        first = text(row[0])
        if not first:
            continue
        if text(row[1]) == "生命":
            group = first
            continue

        name = re.sub(r"\s+", "", first)
        recovered_fields = []
        data_incomplete = False

        def enemy_value(index, label):
            nonlocal data_incomplete
            restored = undate(cells[index])
            if restored is not None:
                recovered_fields.append(label)
                return restored
            value = clean(row[index])
            if isinstance(value, str) and value.strip() in ("待测", "待补充", "数据待补充", "未知", "N/A", "n/a"):
                data_incomplete = True
                return None
            return value

        hp = enemy_value(1, "生命")
        dr = enemy_value(2, "减伤")
        phys = enemy_value(3, "物理伤害")
        element = parse_enemy_element(enemy_value(4, "元素伤害"))
        if recovered_fields:
            warnings.append(f"敌人 {name}：{'/'.join(recovered_fields)}在 Excel 中被存为日期，已按单元格显示格式还原")

        note = text(note_at(cells[0].row))
        raw_location = text(row[6]) or text(location_at(cells[0].row)) or ""
        locations = [p.strip() for p in re.split(r"[;；\n]", raw_location) if p.strip()]
        if not locations and group:
            locations = [group]
        if not locations:
            warnings.append(f"敌人 {name}：没有出现地点")

        out.append({
            "id": make_id(name),
            "name": name,
            "group": group,
            "groupNote": note,
            "groupNoteBlocks": parse_group_note(note),
            "quality": quality_from_fill(cells[0]),
            "hp": hp,
            "damageReduction": dr,
            "restoredFromDate": True if recovered_fields else None,
            "physicalDamage": phys,
            "elementDamage": element,
            "note": text(row[5]),
            "locations": locations,
            "dataIncomplete": True if data_incomplete else None,
        })
    return out


def link_material_entities(materials, catalog):
    """
    有些「材料」本身就是站内的条目，例如皮驮篮既是驮篮也是升级材料。
    给这类材料标记它对应的板块与 id，页面上可以直接跳转过去。
    """
    for m in materials:
        hit = catalog.get(m["name"])
        if hit:
            m["entity"] = {"cat": hit[0], "id": hit[1]}


def link_recipe_entities(consumables, materials, catalog):
    """
    把食物药剂配方里的原料名解析成站内条目引用，供详情页渲染图标框和跳转链接。

    原料横跨多个板块：「生肉」「蘑菇汤」是食物本身，「烈酒」「布料」是材料，
    理论上还可能出现武器护甲。按 食物药剂 → 材料 → 其他板块 的顺序查名字，
    命中就在该原料上写 ref = {"cat", "id", "quality"}。

    quality 带出来是为了让前端画品质色外框。
    查不到的原料保持纯文本，前端渲染成不可点击的普通标签。
    """
    index = {}
    for item in consumables:                       # 食物药剂优先，它们有自己的详情页
        index.setdefault(item["name"], ("consumables", item["id"], item.get("quality")))
    for item in materials:
        index.setdefault(item["name"], ("materials", item["id"], item.get("quality")))
    for name, hit in catalog.items():
        index.setdefault(name, (hit[0], hit[1], None))

    unresolved = set()
    for item in consumables:
        for recipe in item.get("recipes") or []:
            for ingredient in recipe["items"]:
                name = MATERIAL_ALIAS.get(ingredient["name"], ingredient["name"])
                hit = index.get(name)
                if not hit:
                    unresolved.add(ingredient["name"])
                    continue
                ingredient["ref"] = {"cat": hit[0], "id": hit[1], "quality": hit[2]}

    if unresolved:
        warnings.append(
            "食物药剂配方里有 %d 种原料在站内找不到对应条目，将显示为纯文本：%s"
            % (len(unresolved), "、".join(sorted(unresolved)))
        )


# 磨尖等级 → 品质。0/1/2 普通，3/4 稀有，5 独特。
# 这是固定规则，不从单元格底色读，Excel 不需要刷颜色。
SHARPEN_QUALITY = {0: "common", 1: "common", 2: "common",
                   3: "rare", 4: "rare", 5: "unique"}

QUALITY_RANK = {"common": 0, "rare": 1, "unique": 2, "legendary": 3}


def sharpen_quality(base_quality, level):
    """
    磨尖后的品质取「武器自身品质」与「等级对应品质」中更高的一档。

    普通武器基础是 common，结果就是 SHARPEN_QUALITY 的原始规则。
    黑武器基础是 rare，未磨到磨4 都保持蓝框，磨5 才升到独特——
    不会出现一把黑戟未磨时显示白框的情况。
    """
    by_level = SHARPEN_QUALITY[level]
    base = base_quality or "common"
    return base if QUALITY_RANK.get(base, 0) >= QUALITY_RANK[by_level] else by_level

def build_sharpen(weapons):
    """
    磨尖武器表列序：
      0 名称 / 1 0级 / 2 +1 / 3 +2 / 4 +3 / 5 +4 / 6 +5

    第 1 行是跨列标题，第 2 行是表头，从第 3 行起是数据。

    每件武器展开成 6 个磨尖等级，共用同一张武器图标，
    只靠外框品质色和角标区分。品质由等级推导，见 SHARPEN_QUALITY。

    weapons 传入是为了把名称解析成武器 id，图标路径和详情页链接都靠它。
    名称对不上就警告，不静默跳过。
    """
    by_name = {w["name"]: w for w in weapons}
    out = []

    for cells in rows_of("sharpen")[2:]:
        row = list(cells)
        name = text(row[0])
        # 表末尾有几行说明文字，它们没有伤害数值，据此跳过
        if not name or num(row[1]) is None:
            continue

        weapon = by_name.get(name)
        if not weapon:
            warnings.append(f"磨尖武器 {name}：在武器表里找不到同名条目，无法关联图标与详情页")

        base_quality = weapon["quality"] if weapon else "common"
        levels = []
        for lvl in range(6):
            dmg = num(row[1 + lvl])
            if dmg is None:
                warnings.append(f"磨尖武器 {name}：磨{lvl} 没有伤害数值")
            levels.append({
                "level": lvl,
                "damage": dmg,
                "quality": sharpen_quality(base_quality, lvl),
            })

        base = levels[0]["damage"]
        top = levels[5]["damage"]
        out.append({
            "id": make_id(name),
            "name": name,
            "weaponId": weapon["id"] if weapon else None,
            "quality": weapon["quality"] if weapon else "common",
            # 磨尖武器没有自己的图标目录，复用武器图。
            # 图标路径默认按 /images/<板块>/<id>.webp 拼接，这两个字段用来覆盖它。
            "iconCat": "weapons",
            "iconId": weapon["id"] if weapon else None,
            "levels": levels,
            # 列表页直接用的扁平字段，省得前端从 levels 里挑
            "damage0": levels[0]["damage"],
            "damage1": levels[1]["damage"],
            "damage2": levels[2]["damage"],
            "damage3": levels[3]["damage"],
            "damage4": levels[4]["damage"],
            "damage5": levels[5]["damage"],
            "gain": (top - base) if (base is not None and top is not None) else None,
        })

    return out


# 「弯刀（+0/+1/+2）」这类写法，括号里是该盒子能开出的磨尖等级
BOX_DROP_RE = re.compile(r"^(?P<name>.+?)[（(](?P<levels>[+\d/、,，\s]+)[）)]\s*$")

# 盒子品质从名称里的品质词推导，不读单元格底色。
# 盒子表自带一套色板，登记进 QUALITY_BY_FILL 只会让色值表越来越杂，
# 而名称本身已经写明了品质，没有歧义。
BOX_QUALITY_BY_NAME = {"普通": "common", "稀有": "rare",
                       "独特": "unique", "传奇": "legendary", "传说": "legendary"}


def box_quality(name):
    for word, quality in BOX_QUALITY_BY_NAME.items():
        if name.startswith(word):
            return quality
    warnings.append(f"武器盒子 {name}：名称里没有品质词，按普通处理")
    return "common"


def build_boxes(weapons, sharpen=()):
    """
    武器盒子表列序：
      0 盒子名称（合并单元格，只有该组第一行有值） / 1 掉落武器

    第 1 行是标题，第 2 行是表头，从第 3 行起是数据。

    掉落写法两种：
      「配重锤」            —— 固定掉落，无磨尖等级
      「弯刀（+0/+1/+2）」   —— 该盒子能开出未磨、磨1、磨2 三种，拆成三条

    卡片品质取盒子品质：普通盒子开出的都算普通，稀有盒子开出的算稀有，
    以此类推。这和磨尖等级推导出的品质是吻合的（+0~2 普通、+3~4 稀有、+5 独特）。

    weapons 传入用于把名称解析成武器 id，图标与详情页链接都靠它。

    sharpen 传入磨尖武器数据。带磨尖等级、且该武器确实在磨尖表里的掉落，
    链接指向磨尖详情页；否则指向普通武器页。磨尖表目前只收了 23 把，
    盒子里的青铜矛就不在其中，硬跳会 404。
    """
    by_name = {w["name"]: w for w in weapons}
    sharpen_ids = {x["name"]: x["id"] for x in sharpen}
    boxes = []
    current = None

    for row in rows_of("boxes")[2:]:
        row = list(row)
        box_name = text(row[0])
        drop = text(row[1])

        if box_name:                       # 合并单元格只有首行有值
            # make_id 用全局 used_ids 去重，同一个名字调两次会得到 xxx-2。
            # 所以只算一次，iconId 直接复用。
            box_id = make_id(box_name)
            current = {
                "id": box_id,
                "name": box_name,
                "quality": box_quality(box_name),
                "iconCat": "boxes",
                "iconId": box_id,
                "drops": [],
            }
            boxes.append(current)

        if not drop or current is None:
            continue

        matched = BOX_DROP_RE.match(drop)
        if matched:
            base = matched.group("name").strip()
            levels = [int(x.strip().lstrip("+")) for x in
                      re.split(r"[/、,，]", matched.group("levels")) if x.strip()]
        else:
            base, levels = drop, [None]

        weapon = by_name.get(base)
        if not weapon:
            warnings.append(f"武器盒子 {current['name']}：掉落「{base}」在武器表里找不到")

        sharpen_id = sharpen_ids.get(base)
        for lvl in levels:
            # 有磨尖等级且磨尖表里收录了这把武器，才跳磨尖详情页
            if lvl is not None and sharpen_id:
                href = f"/sharpen/{sharpen_id}"
            elif weapon:
                href = f"/weapons/{weapon['id']}"
            else:
                href = None
            current["drops"].append({
                "name": base,
                "weaponId": weapon["id"] if weapon else None,
                "level": lvl,
                "quality": current["quality"],
                "href": href,
                "iconCat": "weapons",
                "iconId": weapon["id"] if weapon else None,
            })

    return boxes


def build_materials(*datasets, skip_names=()):
    """
    从所有配方里反推材料总表。

    skip_names 里的名字不计入。食物药剂的配方原料大半本身就是食物
    （生肉、韭葱、苦药酒…），它们已经有自己的详情页，再塞进材料表
    会让材料列表变得混杂。只有那些在站内没有任何条目的原料
    （烈酒、小空瓶、各类种子）才需要补进来。
    """
    counts = {}
    skip = set(skip_names)
    for data in datasets:
        for item in data:
            for c in item.get("cost", []) or []:
                counts[c["material"]] = counts.get(c["material"], 0) + 1
            for p in item.get("pieces", []) or []:
                for c in p.get("cost", []) or []:
                    counts[c["material"]] = counts.get(c["material"], 0) + 1
            for level in item.get("levels", []) or []:
                for c in level.get("cost", []) or []:
                    counts[c["material"]] = counts.get(c["material"], 0) + 1
            for cost_key in ("buildCost", "assembleCost"):
                for c in item.get(cost_key, []) or []:
                    counts[c["material"]] = counts.get(c["material"], 0) + 1
            # 食物药剂的配方结构不同：recipes[].items[]，字段是 name 不是 material
            for r in item.get("recipes", []) or []:
                for ing in r.get("items", []) or []:
                    name = MATERIAL_ALIAS.get(ing["name"], ing["name"])
                    if name in skip:
                        continue
                    counts[name] = counts.get(name, 0) + 1
    return [
        {"id": make_id(name, "mat"), "name": name, "usedIn": n, "quality": "common"}
        for name, n in sorted(counts.items(), key=lambda kv: -kv[1])
    ]


def normalize_material_name(value):
    """按材料配方的既有规则清理名称并归一化别名。"""
    name = text(value)
    if not name:
        return None
    name = re.sub(r"\s+", "", name)
    return MATERIAL_ALIAS.get(name, name)


def attach_material_descriptions(materials):
    """从可选的材料介绍表给已生成的材料条目附加 description。"""
    def report():
        filled = sum(1 for material in materials if material.get("description"))
        print(f"材料介绍导入：{filled} 条已填写，{len(materials) - filled} 条尚未填写。")

    path = SRC / "材料数据.xlsx"
    if not path.exists():
        print("提示：未找到 data-source/材料数据.xlsx，跳过材料介绍导入。")
        report()
        return

    try:
        rows = cells_of("materials")
    except KeyError:
        warnings.append("材料介绍表缺少工作表「材料介绍」，已跳过材料介绍导入")
        report()
        return

    by_name = {
        normalize_material_name(material["name"]): material
        for material in materials
    }
    for cells in rows[1:]:
        if not cells:
            continue
        name_cell = cells[0]
        name = normalize_material_name(name_cell.value)
        if not name:
            continue
        quality = quality_from_fill(name_cell)
        material = by_name.get(name)
        if material is None:
            warnings.append(f"材料介绍表中的材料「{name}」找不到对应材料，已跳过")
            continue
        material["quality"] = quality
        description = text(cells[1].value if len(cells) > 1 else None)
        if description:
            material["description"] = description

    report()


def link_materials(materials, *datasets):
    """把配方里的材料中文名替换成 material id。"""
    by_name = {m["name"]: m["id"] for m in materials}
    def fix(cost):
        for c in cost or []:
            c["material"] = by_name.get(c["material"], c["material"])
    for data in datasets:
        for item in data:
            fix(item.get("cost"))
            fix(item.get("buildCost"))
            fix(item.get("assembleCost"))
            for p in item.get("pieces", []) or []:
                fix(p.get("cost"))
            for level in item.get("levels", []) or []:
                fix(level.get("cost"))


def write(name, data):
    path = OUT / f"{name}.json"
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"  {path.relative_to(ROOT)}  —  {len(data)} 条")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    print("正在转换…\n")

    weapons = build_weapons()
    armor_sets, armor_pieces = build_armor()
    shields = build_shields()
    backpacks = build_backpacks()
    enemies = build_enemies()
    note_icons = collect_note_icons(enemies)
    amulets = build_amulets()
    scrolls = build_scrolls()
    runes = build_runes()
    consumables = build_consumables()
    sharpen = build_sharpen(weapons)
    boxes = build_boxes(weapons, sharpen)
    cabinets = build_cabinets()
    surface_chests = build_surface_chests()
    fixed_buildings = build_fixed_buildings()
    skills = build_skills()

    materials = build_materials(weapons, armor_sets, armor_pieces, shields,
                                backpacks, amulets, cabinets, surface_chests, fixed_buildings, consumables,
                                skip_names={c["name"] for c in consumables})
    attach_material_descriptions(materials)

    # 材料名若与某个真实条目同名，记录跳转目标
    catalog = {}
    for cat, data in [("weapons", weapons), ("armor", armor_sets),
                      ("armor-pieces", armor_pieces), ("shields", shields),
                      ("backpacks", backpacks), ("surface-chests", surface_chests),
                      ("fixed-buildings", fixed_buildings),
                      ("consumables", consumables)]:
        for it in data:
            catalog.setdefault(it["name"], (cat, it["id"]))
            for pc in it.get("pieces", []) or []:
                catalog.setdefault(pc["name"], ("armor-pieces", pc["id"]))
    link_material_entities(materials, catalog)

    link_materials(materials, weapons, armor_sets, armor_pieces, shields, backpacks, amulets, cabinets, surface_chests, fixed_buildings)

    # 食物药剂的配方原料横跨食物、材料和其他板块，解析成可跳转的引用
    link_recipe_entities(consumables, materials, catalog)

    write("weapons", weapons)
    write("armor", armor_sets)
    write("armor-pieces", armor_pieces)
    write("shields", shields)
    write("backpacks", backpacks)
    write("enemies", enemies)
    for _cat in ("buffs", "debuffs", "enemy-buffs"):
        write(_cat, note_icons.get(_cat, []))
    write("amulets", amulets)
    write("scrolls", scrolls)
    write("runes", runes)
    write("consumables", consumables)
    write("sharpen", sharpen)
    write("boxes", boxes)
    write("cabinets", cabinets)
    write("surface-chests", surface_chests)
    write("fixed-buildings", fixed_buildings)
    write("skills", skills)
    write("materials", materials)

    total_pieces = sum(len(s["pieces"]) for s in armor_sets)
    print(f"\n护甲套装 {len(armor_sets)} 套，含部件 {total_pieces} 件；散件 {len(armor_pieces)} 件")
    parsed_skill_count = sum(1 for skill in skills if not skill["dropLocationsRaw"] or skill["dropLocationsByLevel"])
    print(f"技能掉落地点结构化解析：{parsed_skill_count} 条技能成功，{skill_drop_parse_warnings} 条片段警告")
    print(f"技能「{HIGH_LEVEL_SOURCE_NAME}」规则：{high_level_source_added_count} 条技能已附加，{high_level_source_skipped_count} 条最高只到 Lv10，已跳过")

    if warnings:
        print(f"\n⚠ {len(warnings)} 条需要留意：\n")
        for w in warnings:
            print("  ·", w)
    else:
        print("\n无警告。")


if __name__ == "__main__":
    main()
